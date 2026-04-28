"""triage/triage.py — Quality analysis, document classification, and routing.

Reads all JSON files from output_json/ and produces:
  - triage_report.xlsx   : visual quality report for human review
  - triage_control.json  : routing decisions consumed by the normalizer
  - section_freq.json    : section title frequency across the corpus

Usage:
    python -m triage.triage
    python -m triage.triage --input output_json/ --min-score 70
"""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm


# ─── Language detection ───────────────────────────────────────────────────────

_LANG_NAMES = {
    "english", "français", "french", "deutsch", "german", "español", "spanish",
    "italiano", "italian", "português", "portugese", "portuguese",
    "nederlands", "dutch", "svenska", "swedish", "dansk", "danish",
    "polski", "polish", "magyar", "hungarian", "čeština", "czech",
    "русский", "russian", "українська", "ukrainian", "arabic", "hebrew",
    "suomi", "finnish", "norsk", "norwegian", "eesti", "estonian",
    "中文", "traditional chinese", "simplified chinese",
}


def _count_language_headers(sections: list[dict]) -> int:
    count = 0
    for s in sections:
        title = s.get("raw_title", "").lower().strip()
        parts = re.split(r"[/|\\]", title)
        for part in parts:
            if part.strip() in _LANG_NAMES:
                count += 1
                break
    return count


# ─── Problem detection helpers ────────────────────────────────────────────────

_STRUCTURAL_TITLES = {"UNTITLED", "OPENING", "CLOSING"}
_STRUCTURAL_NORMS  = {"untitled", "opening", "closing", "other"}


def is_code_header(text: str) -> bool:
    text = text.strip()
    return bool(
        re.match(r"^[A-Z]{2,5}-\d{2,}$", text) or
        re.match(r"^[A-Z]-[A-Z0-9]{4,}$", text) or
        re.match(r"^[A-Z]{1,4}\d{2,}[A-Z0-9\-]*$", text)
    )


def is_list_header(text: str) -> bool:
    text = text.strip().lower()
    return bool(re.match(r"^[a-z]\)", text) or re.match(r"^\d+\.$", text))


def is_long_header(text: str) -> bool:
    return len(text.split()) > 12


def is_very_long_header(text: str) -> bool:
    return len(text.split()) > 18


def is_footerish_cell(text: str) -> bool:
    t = (text or "").strip()
    if not t:
        return False
    if re.match(r"^[A-Z0-9\-]{6,}$", t):
        return True
    if re.match(r"^[ivxlcdm]+$", t.lower()):
        return True
    if re.match(r"^\d+$", t):
        return True
    return False


def analyze_tables(tables: list[dict]) -> dict:
    suspicious = footerish = 0
    for table in tables:
        columns = table.get("columns", []) or []
        rows    = table.get("rows", []) or []
        s = f = False
        if len(columns) > 6:
            s = True
        for row in rows:
            if len(row) > 8:
                s = True
            for cell in row:
                if len(str(cell)) > 120:
                    s = True
                if is_footerish_cell(str(cell)):
                    f = True
        for col in columns:
            if len(str(col)) > 120:
                s = True
            if is_footerish_cell(str(col)):
                f = True
        if s:
            suspicious += 1
        if f:
            footerish += 1
    return {"suspicious_tables": suspicious, "footer_tables": footerish}


def count_headers(pages: list[dict]) -> tuple[list[str], list[str]]:
    section_headers, subheaders = [], []
    for page in pages:
        for h in page.get("header_candidates", []):
            level = h.get("header_level", "")
            text  = (h.get("text") or "").strip()
            if not text:
                continue
            if level == "section_header":
                section_headers.append(text)
            elif level == "subheader":
                subheaders.append(text)
    return section_headers, subheaders


def detect_problems(
    page_count: int,
    sections: list[dict],
    pages: list[dict],
    tables: list[dict],
) -> tuple[list[str], int]:
    problems = []
    total_lines  = sum(p.get("text_line_count", 0) for p in pages)
    total_images = sum(p.get("image_count", 0) for p in pages)
    section_count = len(sections)

    untitled      = sum(1 for s in sections if s.get("raw_title") in _STRUCTURAL_TITLES)
    other_count   = sum(1 for s in sections if s.get("normalized_title") == "other")
    other_ratio   = (other_count / section_count) if section_count else 0

    section_headers, subheaders = count_headers(pages)
    list_hdrs      = sum(1 for h in section_headers if is_list_header(h))
    code_hdrs      = sum(1 for h in section_headers if is_code_header(h))
    long_hdrs      = sum(1 for h in section_headers if is_long_header(h))
    very_long_hdrs = sum(1 for h in section_headers if is_very_long_header(h))

    sec_sizes = [
        len(s.get("text_plain", ""))
        for s in sections if s.get("raw_title") not in _STRUCTURAL_TITLES
    ]
    avg_sec_size  = int(sum(sec_sizes) / len(sec_sizes)) if sec_sizes else 0
    table_stats   = analyze_tables(tables)

    if total_lines == 0 and page_count > 0:
        problems.append("no_text")
    if total_lines < 15 and total_images >= 5:
        problems.append("mostly_images")
    if (
        (page_count <= 4  and section_count > 8)  or
        (5 <= page_count <= 10 and section_count > 15) or
        (page_count > 10 and section_count > page_count * 2)
    ):
        problems.append("too_many_sections")
    if section_count >= 4 and untitled / section_count > 0.30:
        problems.append("many_untitled")
    if list_hdrs >= 2:
        problems.append("list_headers")
    if code_hdrs >= 1:
        problems.append("code_headers")
    if long_hdrs >= 2 or very_long_hdrs >= 1:
        problems.append("long_headers")
    if section_count >= 8 and other_ratio > 0.60:
        problems.append("poor_normalization")
    if table_stats["suspicious_tables"] >= 1:
        problems.append("suspicious_tables")
    if table_stats["footer_tables"] >= 1:
        problems.append("footer_tables")
    if section_count >= 6 and avg_sec_size < 150:
        problems.append("sections_too_short")
    if section_count <= 4 and avg_sec_size > 3000:
        problems.append("sections_too_long")
    if subheaders and len(subheaders) > max(6, len(section_headers) * 2):
        problems.append("many_subheaders")

    _is_multilingual = False
    if _count_language_headers(sections) >= 3:
        _is_multilingual = True
    elif section_count > 20 and other_ratio > 0.85:
        _is_multilingual = True
    else:
        first_pages_text = [
            h.get("text", "").lower()
            for p in pages[:4]
            for h in p.get("header_candidates", [])
        ]
        if sum(1 for t in first_pages_text if any(lang in t for lang in _LANG_NAMES)) >= 4:
            _is_multilingual = True
    if _is_multilingual:
        problems.append("multilingual")

    penalties = {
        "no_text": 50, "mostly_images": 30,
        "too_many_sections": 12, "many_untitled": 10,
        "list_headers": 8, "code_headers": 8, "long_headers": 6,
        "poor_normalization": 12, "suspicious_tables": 10,
        "footer_tables": 10, "sections_too_short": 6,
        "sections_too_long": 6, "many_subheaders": 5,
    }
    quality_score = max(0, 100 - sum(penalties.get(p, 0) for p in problems))
    return problems, quality_score


# ─── Document classification ──────────────────────────────────────────────────

def classify_doc(page_count: int, section_count: int, table_count: int, image_count: int) -> str:
    if page_count <= 3:
        return "short_doc"
    if page_count > 20 or section_count > 15:
        return "long_doc"
    if table_count >= 5 and section_count <= 10:
        return "datasheet"
    return "medium_doc"


# ─── Helpers ──────────────────────────────────────────────────────────────────

def sanitize(text: str) -> str:
    if not text:
        return ""
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", str(text))


def top_sections(sections: list[dict], n: int = 5) -> str:
    titles = [
        s["raw_title"] for s in sections
        if s["raw_title"] not in _STRUCTURAL_TITLES and len(s["raw_title"]) > 2
    ]
    return " | ".join(titles[:n])


def top_normalized(sections: list[dict], n: int = 8) -> str:
    norm_counts = Counter(
        s["normalized_title"] for s in sections
        if s["normalized_title"] not in _STRUCTURAL_NORMS
    )
    return " | ".join(k for k, _ in norm_counts.most_common(n))


def total_text_length(sections: list[dict]) -> int:
    return sum(len(s.get("text_plain", "")) for s in sections)


# ─── Main analysis ────────────────────────────────────────────────────────────

def analyze_json(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    doc        = data.get("document", {})
    pages      = data.get("pages", [])
    sections   = data.get("sections", [])
    tables     = data.get("tables", [])
    images     = data.get("images", [])
    enrichment = data.get("enrichment", {})

    field_hints = enrichment.get("field_hints", []) or []
    detected_fields = {h.get("target_field") for h in field_hints if h.get("target_field")}

    page_count    = doc.get("page_count", len(pages))
    section_count = len(sections)
    table_count   = len(tables)
    image_count   = len(images)
    text_length   = total_text_length(sections)
    untitled      = sum(1 for s in sections if s["raw_title"] in _STRUCTURAL_TITLES)

    sec_sizes = [
        len(s.get("text_plain", ""))
        for s in sections if s["raw_title"] not in _STRUCTURAL_TITLES
    ]
    avg_sec_size = int(sum(sec_sizes) / len(sec_sizes)) if sec_sizes else 0

    doc_type = classify_doc(page_count, section_count, table_count, image_count)
    problems, quality_score = detect_problems(page_count, sections, pages, tables)

    return {
        # Excel columns
        "filename":               sanitize(doc.get("file_name", path.stem)),
        "doc_type":               doc_type,
        "quality_score":          quality_score,
        "pages":                  page_count,
        "sections":               section_count,
        "untitled_sections":      untitled,
        "tables":                 table_count,
        "images":                 image_count,
        "total_chars":            text_length,
        "avg_chars_per_section":  avg_sec_size,
        "detected_fields":        len(detected_fields),
        "field_hints":            len(field_hints),
        "problems":               sanitize(" | ".join(problems)),
        "top_sections":           sanitize(top_sections(sections)),
        "detected_categories":    sanitize(top_normalized(sections)),
        "title_metadata":         sanitize(doc.get("title_metadata") or ""),
        "author_metadata":        sanitize(doc.get("author_metadata") or ""),
        # Internal — not written to Excel
        "_json_path":             str(path),
        "_sections":              sections,
    }


# ─── Control JSON ─────────────────────────────────────────────────────────────

def write_control_json(rows: list[dict], output_path: Path, min_score: int) -> None:
    """Write triage_control.json with a routing decision per document.

    Each entry:
    {
        "filename":      "Product Manual XYZ.pdf",
        "json_path":     "output_json/Product Manual XYZ.json",
        "doc_type":      "medium_doc",
        "quality_score": 88,
        "problems":      "long_headers | suspicious_tables",
        "action":        "process" | "process_primary_language" | "review" | "skip",
        "reason":        "..."
    }
    """
    control = []
    for row in rows:
        score    = row["quality_score"]
        problems = _problems_set(row)
        filename = row["filename"]

        if problems & {"no_text", "mostly_images"}:
            action = "skip"
            reason = next(p for p in ("no_text", "mostly_images") if p in problems)
        elif "multilingual" in problems:
            action = "process_primary_language"
            reason = "multilingual document — extract primary language sections only"
        elif score < min_score:
            action = "skip"
            reason = f"quality_score {score} below threshold {min_score}"
        else:
            action = "process"
            reason = ""

        control.append({
            "filename":      filename,
            "json_path":     row["_json_path"],
            "doc_type":      row["doc_type"],
            "quality_score": score,
            "problems":      row.get("problems", ""),
            "action":        action,
            "reason":        reason,
        })

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(control, f, ensure_ascii=False, indent=2)

    action_counts = Counter(c["action"] for c in control)
    print(f"Control JSON saved to: {output_path}")
    for action, count in action_counts.most_common():
        print(f"  {action}: {count} documents")


# ─── Section frequency ────────────────────────────────────────────────────────

def build_section_freq(rows: list[dict]) -> list[dict]:
    """Count how often each normalized_title appears across all documents."""
    freq: dict[str, dict] = {}

    for row in rows:
        seen_in_doc: set[str] = set()
        for sec in row.get("_sections", []):
            norm = sec.get("normalized_title") or "other"
            raw  = (sec.get("raw_title") or "").strip()
            tlen = len(sec.get("text_plain", ""))

            if norm not in freq:
                freq[norm] = {
                    "normalized_title": norm,
                    "doc_count":        0,
                    "raw_titles":       Counter(),
                    "text_lens":        [],
                    "is_structural":    norm in _STRUCTURAL_NORMS,
                }

            if raw:
                freq[norm]["raw_titles"][raw] += 1
            freq[norm]["text_lens"].append(tlen)
            seen_in_doc.add(norm)

        for norm in seen_in_doc:
            freq[norm]["doc_count"] += 1

    result = []
    for data in freq.values():
        top_raws  = "; ".join(f"{r} ({c})" for r, c in data["raw_titles"].most_common(6))
        lens      = data["text_lens"]
        avg_chars = int(sum(lens) / len(lens)) if lens else 0
        result.append({
            "normalized_title":  data["normalized_title"],
            "doc_count":         data["doc_count"],
            "unique_raw_titles": len(data["raw_titles"]),
            "avg_chars":         avg_chars,
            "is_structural":     data["is_structural"],
            "raw_variations":    sanitize(top_raws),
        })

    result.sort(key=lambda x: (-x["doc_count"], -x["unique_raw_titles"]))
    return result


def write_section_freq_sheet(wb, freq_rows: list[dict]) -> None:
    """Add the 'Section Frequency' sheet to an existing workbook."""
    ws = wb.create_sheet("Section Frequency")

    FREQ_HEADERS = [
        ("normalized_title",  24),
        ("doc_count",         16),
        ("unique_raw_titles", 16),
        ("avg_chars",         12),
        ("is_structural",     14),
        ("raw_variations",    80),
    ]

    total_docs = max((r["doc_count"] for r in freq_rows), default=1)

    def _freq_fill(docs: int, structural: bool) -> PatternFill:
        if structural:
            return PatternFill("solid", start_color="E8E8E8")
        ratio = docs / total_docs
        if ratio >= 0.75:
            return PatternFill("solid", start_color="C6EFCE")
        if ratio >= 0.40:
            return PatternFill("solid", start_color="D6E4F7")
        if ratio >= 0.15:
            return PatternFill("solid", start_color="FFF3CD")
        return PatternFill("solid", start_color="FADADD")

    for col_idx, (name, width) in enumerate(FREQ_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(freq_rows, start=2):
        fill = _freq_fill(row["doc_count"], row["is_structural"])
        for col_idx, (col_name, _) in enumerate(FREQ_HEADERS, start=1):
            val  = row.get(col_name, "")
            cell = ws.cell(
                row=row_idx, column=col_idx,
                value=sanitize(val) if isinstance(val, str) else val,
            )
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_name == "raw_variations"))
        ws.row_dimensions[row_idx].height = 18

    ws.auto_filter.ref = f"A1:{get_column_letter(len(FREQ_HEADERS))}1"

    legend_row = len(freq_rows) + 3
    ws.cell(row=legend_row, column=1, value="Legend").font = Font(bold=True, name="Arial", size=9)
    legend_items = [
        ("C6EFCE", "≥75% of docs — core corpus section"),
        ("D6E4F7", "40–74% of docs — frequent section"),
        ("FFF3CD", "15–39% of docs — occasional section"),
        ("FADADD", "<15% of docs — rare / domain-specific section"),
        ("E8E8E8", "Structural (opening / closing / untitled)"),
    ]
    for i, (color, desc) in enumerate(legend_items, start=legend_row + 1):
        ws.cell(row=i, column=1, value="").fill = PatternFill("solid", start_color=color)
        ws.cell(row=i, column=2, value=desc).font = Font(name="Arial", size=9)


# ─── Excel ────────────────────────────────────────────────────────────────────

HEADERS = [
    ("filename",              42),
    ("doc_type",              18),
    ("quality_score",         12),
    ("pages",                  8),
    ("sections",              10),
    ("untitled_sections",     16),
    ("tables",                 8),
    ("images",                 8),
    ("total_chars",           13),
    ("avg_chars_per_section", 18),
    ("detected_fields",       16),
    ("field_hints",           12),
    ("problems",              50),
    ("top_sections",          60),
    ("detected_categories",   50),
    ("title_metadata",        30),
    ("author_metadata",       25),
]

TYPE_COLORS = {
    "long_doc":   "D6E4F7",
    "medium_doc": "D6F7DC",
    "short_doc":  "FFF3CD",
    "datasheet":  "F7D6F0",
}

HEADER_FILL = PatternFill("solid", start_color="2B4A8C")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT   = Font(name="Arial", size=9)
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _problems_set(row: dict) -> set[str]:
    raw = row.get("problems", "") or ""
    return {p.strip() for p in raw.split("|") if p.strip()}


def _row_fill(row: dict) -> PatternFill:
    score    = row.get("quality_score", 100)
    problems = _problems_set(row)
    if score < 50 or problems & {"no_text", "mostly_images"}:
        return PatternFill("solid", start_color="FADADD")
    if score < 75 or "multilingual" in problems:
        return PatternFill("solid", start_color="FFF3CD")
    color = TYPE_COLORS.get(row.get("doc_type", ""), "FFFFFF")
    return PatternFill("solid", start_color=color)


def write_excel(rows: list[dict], output_path: Path, freq_rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Triage"

    for col_idx, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, start=2):
        fill = _row_fill(row)
        for col_idx, (col_name, _) in enumerate(HEADERS, start=1):
            val  = row.get(col_name, "")
            cell = ws.cell(
                row=row_idx, column=col_idx,
                value=sanitize(val) if isinstance(val, str) else val,
            )
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx >= 11))
        ws.row_dimensions[row_idx].height = 18

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    ws2 = wb.create_sheet("Legend")
    ws2["A1"] = "Color / Type"
    ws2["B1"] = "Meaning"
    ws2["A1"].font = Font(bold=True, name="Arial")
    ws2["B1"].font = Font(bold=True, name="Arial")
    legend = [
        ("FADADD", "Critical — quality < 50 or no text: skipped in normalization"),
        ("FFF3CD", "Review — quality 50–74 or multilingual: special processing"),
        ("D6E4F7", "Long document — >20 pages or >15 sections"),
        ("D6F7DC", "Medium document — 4–20 pages"),
        ("F7D6F0", "Datasheet — high table density"),
    ]
    for i, (color, desc) in enumerate(legend, start=2):
        ws2.cell(row=i, column=1, value="").fill = PatternFill("solid", start_color=color)
        ws2.cell(row=i, column=2, value=desc).font = Font(name="Arial", size=9)
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 65

    write_section_freq_sheet(wb, freq_rows)
    wb.save(output_path)
    print(f"Excel saved to: {output_path}")


# ─── Public API ───────────────────────────────────────────────────────────────

def run_triage(
    input_dir:    Path | str = "output_json",
    output_path:  Path | str = "triage_report.xlsx",
    control_path: Path | str = "triage_control.json",
    min_score:    int = 50,
) -> None:
    input_dir    = Path(input_dir)
    output_path  = Path(output_path)
    control_path = Path(control_path)

    json_files = sorted(input_dir.glob("*.json"))
    if not json_files:
        print(f"No JSON files found in {input_dir.resolve()}")
        return

    print(f"Analyzing {len(json_files)} files... (quality threshold: {min_score})")
    rows = []
    for path in tqdm(json_files, desc="Triaging", unit="file"):
        try:
            rows.append(analyze_json(path))
            print(f"  ✓ {path.name}")
        except Exception as e:
            print(f"  ✗ {path.name}: {e}")

    rows.sort(key=lambda r: (-r["quality_score"], r["doc_type"]))
    freq_rows = build_section_freq(rows)

    freq_path = control_path.parent / "section_freq.json"
    with open(freq_path, "w", encoding="utf-8") as f:
        json.dump({"total_docs": len(rows), "sections": freq_rows}, f, ensure_ascii=False, indent=2)
    print(f"Section frequency saved to: {freq_path}")

    write_excel(rows, output_path, freq_rows)
    print()
    write_control_json(rows, control_path, min_score=min_score)

    print(f"\nBy type:")
    for doc_type, count in Counter(r["doc_type"] for r in rows).most_common():
        print(f"  {doc_type}: {count}")
    print(f"\nBy quality:")
    for lo, hi, label in [(0, 50, "critical"), (50, 75, "low"), (75, 90, "medium"), (90, 101, "high")]:
        n = sum(1 for r in rows if lo <= r["quality_score"] < hi)
        print(f"  {label} ({lo}–{hi-1}): {n}")


# ─── Entry point ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Triage extractor JSON files for normalization.")
    parser.add_argument("--input",     default="output_json",         help="Folder containing extractor JSON files")
    parser.add_argument("--output",    default="triage_report.xlsx",  help="Output Excel report path")
    parser.add_argument("--control",   default="triage_control.json", help="Output control JSON for the normalizer")
    parser.add_argument("--min-score", default=60, type=int,          help="Minimum quality score to process (default: 60)")
    args = parser.parse_args()

    run_triage(
        input_dir=args.input,
        output_path=args.output,
        control_path=args.control,
        min_score=args.min_score,
    )


if __name__ == "__main__":
    main()