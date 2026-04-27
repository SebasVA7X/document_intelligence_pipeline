"""
normalizer/exporter.py — Converts the list of normalized rows to Excel.

Workbook structure:
  - Short   : short documents  (doc_type == "short_doc")
  - Medium  : medium documents (doc_type == "medium_doc")
  - Long    : long documents   (doc_type == "long_doc")
  - Skipped : documents skipped by triage
  - Columns : fixed identification fields + dynamic section columns
  - One row per document; full section text in each cell
"""
from __future__ import annotations

import re as _re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Characters illegal in Excel cells (rejected by openpyxl)
_ILLEGAL_CHARS = _re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\ufffe\uffff]")


def _sanitize(value: Any) -> Any:
    """Strip illegal characters from strings before writing to Excel."""
    if isinstance(value, str):
        return _ILLEGAL_CHARS.sub("", value)
    return value


# ─── Fixed identification columns ────────────────────────────────────────────
ID_COLUMNS: list[tuple[str, int]] = [
    ("filename",         45),
    ("code",              15),
    ("product_name",      35),
    ("doc_type",        12),
    ("pages",              7),
    ("quality_score",   10),
]

# Width for content columns (long text)
CONTENT_WIDTH = 60

# ─── Styles ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="2B4A8C")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
BODY_FONT   = Font(name="Arial", size=8)
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── tipo_doc → sheet name and row fill colour ────────────────────────────────
_SHEET_CONFIG: dict[str, tuple[str, str]] = {
    "short_doc":    ("Short",    "FFF3CD"),
    "medium_doc":   ("Medium",   "D6F7DC"),
    "long_doc":     ("Long",     "D6E4F7"),
    "skipped":      ("Skipped",  "F0F0F0"),
}

_TIPO_ORDER = ["medium_doc", "short_doc", "long_doc", "skipped"]


def _col_headers(columns: list[str]) -> list[tuple[str, int]]:
    """Full column list (id + dynamic sections) with widths."""
    section_cols = [(col, CONTENT_WIDTH) for col in columns]
    return ID_COLUMNS + section_cols


def _write_sheet(ws, rows: list[dict[str, Any]], tipo: str, columns: list[str]) -> None:
    _, fill_color = _SHEET_CONFIG.get(tipo, ("", "FFFFFF"))
    headers = _col_headers(columns)
    columns_set = set(columns)

    for col_idx, (name, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    col_names = [name for name, _ in headers]
    row_fill  = PatternFill("solid", start_color=fill_color)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(col_names, start=1):
            val  = row.get(col_name, "")
            cell = ws.cell(
                row=row_idx, column=col_idx,
                value=_sanitize(val),
            )
            cell.font      = BODY_FONT
            cell.fill      = row_fill
            cell.border    = BORDER
            is_content = col_name in columns_set
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=is_content,
                horizontal="left" if is_content else "center",
            )
        ws.row_dimensions[row_idx].height = 60

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def rows_to_excel(rows: list[dict[str, Any]], output_path: Path, columns: list[str]) -> None:
    """Generate the final Excel matrix.

    Args:
        rows:        List of rows (one per document), output of mapper.build_row.
        output_path: Output .xlsx file path.
        columns:     Dynamic column list (determined by router).
    """
    groups: dict[str, list[dict]] = {t: [] for t in _TIPO_ORDER}
    for row in rows:
        doc_type = row.get("doc_type", "")
        action   = row.get("action", "")
        if action == "skip":
            groups["skipped"].append(row)
        elif doc_type in groups:
            groups[doc_type].append(row)
        else:
            groups["medium_doc"].append(row)

    wb = Workbook()
    wb.remove(wb.active)

    for tipo in _TIPO_ORDER:
        grupo = groups[tipo]
        if not grupo:
            continue
        sheet_title, _ = _SHEET_CONFIG[tipo]
        ws = wb.create_sheet(title=sheet_title)
        _write_sheet(ws, grupo, tipo, columns)
        print(f"  Sheet '{sheet_title}': {len(grupo)} documents")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Output saved to: {output_path}")
